import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "2026-06-11 开发内容"

headers = ["#", "时间", "Commit", "类型", "功能模块", "详细描述", "涉及文件数", "影响范围"]
ws.append(headers)

header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="D9DEE7"),
    right=Side(style="thin", color="D9DEE7"),
    top=Side(style="thin", color="D9DEE7"),
    bottom=Side(style="thin", color="D9DEE7"),
)

for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border

data = [
    [1, "10:19", "15519cc9", "修复", "聊天输入", "IME输入法按回车误发送 — 用ref手动追踪composition状态，避免中文输入时回车直接发送消息", "3", "TopBarChatInput.tsx"],
    [2, "14:57", "fb93443d", "重构", "AI聊天架构", "AI聊天架构统一 — ChatPage薄包装+硬编码URL修复+Session存储统一。将ChatPage改为薄包装组件，CrossWmsChat成为主组件", "8", "ChatPage, CrossWmsChat, sessionStore"],
    [3, "18:45", "5dceebfc", "修复", "构建脚本", "构建脚本从git remote动态获取GITHUB_OWNER/GITHUB_REPO，修复硬编码cdf-know-clow导致Release创建失败", "2", "build-dmg-pywebview.sh"],
    [4, "19:59", "69542783", "CI", "GitHub Actions", "修复并完善GitHub Actions CI/CD工作流", "1", ".github/workflows"],
    [5, "20:02", "69e8892d", "修复", "CI", "修复npm ci peer dependency + release notes变量名", "1", ".github/workflows"],
    [6, "20:05", "ba19704b", "修复", "CI", "Fetch tags加容错，避免tag push场景报错", "1", ".github/workflows"],
    [7, "20:09", "c9fac74e", "修复", "CI", "添加permissions: contents: write解决Release创建403错误", "1", ".github/workflows"],
    [8, "20:21", "2419c715", "功能", "聊天组件", "v1.5.15: 删除输入时智能提示 + 滚动条修复 + 新聊天组件。移除SkillSuggestionPopover和SkillSelector，简化输入框", "12", "CrossWmsChat, TopBarChatInput"],
]

zebra1 = PatternFill(fill_type="solid", fgColor="FFFFFF")
zebra2 = PatternFill(fill_type="solid", fgColor="F7F9FC")

for i, row_data in enumerate(data, start=2):
    ws.append(row_data)
    fill = zebra1 if (i - 2) % 2 == 0 else zebra2
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=i, column=col)
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if col == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 55
ws.column_dimensions["G"].width = 12
ws.column_dimensions["H"].width = 28

ws.row_dimensions[1].height = 28
for i in range(2, len(data) + 2):
    ws.row_dimensions[i].height = 55

ws.freeze_panes = "A2"

ws2 = wb.create_sheet("统计汇总")
ws2.append(["统计项", "数值"])
ws2.append(["总Commit数", "8"])
ws2.append(["功能新增", "2"])
ws2.append(["Bug修复", "4"])
ws2.append(["架构重构", "1"])
ws2.append(["CI/CD优化", "4"])
ws2.append(["新增文件", "~15"])
ws2.append(["修改文件", "~36"])
ws2.append(["删除文件", "~8"])
ws2.append(["代码净增", "+1010行"])

for col in range(1, 3):
    cell = ws2.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border

for row in range(2, 11):
    for col in range(1, 3):
        cell = ws2.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        if row % 2 == 0:
            cell.fill = zebra2

ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 15

wb.save("/sessions/6a26c727a46ded7e9534a54c/workspace/development-list-2026-06-11.xlsx")
print("Saved")
